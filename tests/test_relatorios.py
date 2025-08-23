import csv
import io
from unittest.mock import patch

import pytest
from django.utils import timezone
from rest_framework.exceptions import ValidationError
from rest_framework.test import APIRequestFactory, force_authenticate

from accounts.factories import UserFactory
from accounts.models import UserType
from financeiro.models import CentroCusto, LancamentoFinanceiro
from financeiro.services.relatorios import gerar_relatorio
from financeiro.services.exportacao import exportar_para_arquivo
from financeiro.views.api import FinanceiroViewSet, parse_periodo
from organizacoes.factories import OrganizacaoFactory

pytestmark = pytest.mark.django_db


@pytest.fixture
def api_client():
    from rest_framework.test import APIClient

    return APIClient()


def test_cache_key_multiplos_centros():
    user = UserFactory(user_type=UserType.ADMIN)
    c1 = CentroCusto.objects.create(nome="C1", tipo="organizacao")
    c2 = CentroCusto.objects.create(nome="C2", tipo="organizacao")
    factory = APIRequestFactory()
    view = FinanceiroViewSet.as_view({"get": "relatorios"})
    with patch("financeiro.views.api.gerar_relatorio") as mock_rel:
        mock_rel.return_value = {"saldo_atual": 0, "serie": [], "inadimplencia": [], "total_inadimplentes": 0}
        req1 = factory.get(
            "/api/financeiro/relatorios/", {"centro": [str(c1.id), str(c2.id)]}
        )
        force_authenticate(req1, user=user)
        view(req1)
        req2 = factory.get(
            "/api/financeiro/relatorios/", {"centro": [str(c2.id), str(c1.id)]}
        )
        force_authenticate(req2, user=user)
        view(req2)
    assert mock_rel.call_count == 1


def test_parse_periodo_helpers():
    dt = parse_periodo("2024-05")
    assert dt.year == 2024 and dt.month == 5
    with pytest.raises(ValidationError):
        parse_periodo("2024/05")


def test_tipo_filter_relatorio():
    org = OrganizacaoFactory()
    centro = CentroCusto.objects.create(nome="C", tipo="organizacao", organizacao=org)
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=50,
        tipo=LancamentoFinanceiro.Tipo.APORTE_INTERNO,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PAGO,
    )
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=-20,
        tipo=LancamentoFinanceiro.Tipo.DESPESA,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PAGO,
    )
    data = gerar_relatorio(centro=str(centro.id), tipo="receitas")
    assert data["serie"][0]["despesas"] == 0.0
    data = gerar_relatorio(centro=str(centro.id), tipo="despesas")
    assert data["serie"][0]["receitas"] == 0.0


def test_status_filter_relatorio():
    org = OrganizacaoFactory()
    centro = CentroCusto.objects.create(nome="C", tipo="organizacao", organizacao=org)
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=100,
        tipo=LancamentoFinanceiro.Tipo.APORTE_INTERNO,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PENDENTE,
    )
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=50,
        tipo=LancamentoFinanceiro.Tipo.APORTE_INTERNO,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PAGO,
    )
    data = gerar_relatorio(centro=str(centro.id), status=LancamentoFinanceiro.Status.PAGO)
    assert data["serie"][0]["receitas"] == 50.0
    data = gerar_relatorio(centro=str(centro.id), status=LancamentoFinanceiro.Status.PENDENTE)
    assert data["serie"][0]["receitas"] == 100.0


def test_total_inadimplentes():
    org = OrganizacaoFactory()
    centro = CentroCusto.objects.create(nome="C", tipo="organizacao", organizacao=org)
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=100,
        tipo=LancamentoFinanceiro.Tipo.APORTE_INTERNO,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PENDENTE,
    )
    LancamentoFinanceiro.objects.create(
        centro_custo=centro,
        valor=-30,
        tipo=LancamentoFinanceiro.Tipo.DESPESA,
        data_lancamento=timezone.now(),
        status=LancamentoFinanceiro.Status.PENDENTE,
    )
    data = gerar_relatorio(centro=str(centro.id))
    assert data["total_inadimplentes"] == 100.0


def test_exportacoes():
    linhas = [["1", "2"], ["3", "4"]]
    path = exportar_para_arquivo("csv", ["a", "b"], linhas)
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert rows[0] == ["a", "b"]
    pytest.importorskip("openpyxl")
    path = exportar_para_arquivo("xlsx", ["a", "b"], linhas)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    assert ws[1][0].value == "a"
