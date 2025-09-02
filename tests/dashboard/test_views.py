import io
from pathlib import Path

import pytest
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse

from core.permissions import (
    AdminRequiredMixin,
    ClienteRequiredMixin,
    GerenteRequiredMixin,
    SuperadminRequiredMixin,
)
from dashboard.models import DashboardConfig, DashboardFilter
from dashboard.services import DashboardMetricsService
from discussao.models import CategoriaDiscussao, RespostaDiscussao, TopicoDiscussao
from feed.factories import PostFactory

try:
    import weasyprint  # type: ignore
except Exception:  # pragma: no cover - optional
    weasyprint = None
from dashboard.views import (
    AdminDashboardView,
    ClienteDashboardView,
    DashboardBaseView,
    GerenteDashboardView,
    RootDashboardView,
)

pytestmark = pytest.mark.django_db


def _assert_metrics_in_context(context):
    for key in [
        "num_users",
        "num_organizacoes",
        "num_nucleos",
        "num_empresas",
        "num_eventos",
    ]:
        assert key in context
        assert set(context[key].keys()) == {"total", "crescimento"}


def test_base_view_mixins():
    assert issubclass(DashboardBaseView, LoginRequiredMixin)


def test_view_mixins():
    assert issubclass(RootDashboardView, SuperadminRequiredMixin)
    assert issubclass(AdminDashboardView, AdminRequiredMixin)
    assert issubclass(GerenteDashboardView, GerenteRequiredMixin)
    assert issubclass(ClienteDashboardView, ClienteRequiredMixin)


def test_root_dashboard_view(client, root_user):
    client.force_login(root_user)
    resp = client.get(reverse("dashboard:root"))
    assert resp.status_code == 200
    assert "dashboard/root.html" in [t.name for t in resp.templates]
    _assert_metrics_in_context(resp.context)
    assert "Dashboard Root" in resp.content.decode()


def test_admin_dashboard_view(client, admin_user):
    client.force_login(admin_user)
    resp = client.get(reverse("dashboard:admin"))
    assert resp.status_code == 200
    assert "dashboard/admin.html" in [t.name for t in resp.templates]
    _assert_metrics_in_context(resp.context)
    assert "Dashboard Administrativo" in resp.content.decode()


def test_gerente_dashboard_view(client, gerente_user):
    client.force_login(gerente_user)
    resp = client.get(reverse("dashboard:gerente"))
    assert resp.status_code == 200
    assert "dashboard/gerente.html" in [t.name for t in resp.templates]
    _assert_metrics_in_context(resp.context)
    assert "Dashboard Gerente" in resp.content.decode()


def test_cliente_dashboard_view(client, cliente_user):
    client.force_login(cliente_user)
    resp = client.get(reverse("dashboard:cliente"))
    assert resp.status_code == 200
    assert "dashboard/cliente.html" in [t.name for t in resp.templates]
    _assert_metrics_in_context(resp.context)
    assert "Dashboard Cliente" in resp.content.decode()


def test_base_view_accepts_params(monkeypatch, client, admin_user):
    client.force_login(admin_user)

    def fake_metrics(user, periodo="mensal", inicio=None, fim=None, escopo="auto", **filters):
        assert periodo == "anual"
        assert escopo == "organizacao"
        assert filters["organizacao_id"] == str(admin_user.organizacao_id)
        assert filters["metricas"] == ["num_users"]
        return {
            "num_users": {"total": 1, "crescimento": 0.0},
        }

    monkeypatch.setattr(DashboardMetricsService, "get_metrics", fake_metrics)

    resp = client.get(
        reverse("dashboard:admin"),
        {
            "periodo": "anual",
            "escopo": "organizacao",
            "organizacao_id": admin_user.organizacao_id,
            "metricas": ["num_users"],
        },
    )
    assert resp.status_code == 200
    assert resp.context["periodo"] == "anual"
    assert resp.context["escopo"] == "organizacao"
    assert resp.context["filtros"]["organizacao_id"] == str(admin_user.organizacao_id)
    assert resp.context["metricas_selecionadas"] == ["num_users"]
    assert resp.context["chart_data"] == [1]


def test_invalid_date_returns_400(client, admin_user):
    client.force_login(admin_user)
    resp = client.get(reverse("dashboard:admin"), {"data_inicio": "xx"})
    assert resp.status_code == 400


def test_metrics_partial_new_metrics(client, admin_user):
    client.force_login(admin_user)
    PostFactory(autor=admin_user, organizacao=admin_user.organizacao)
    cat = CategoriaDiscussao.objects.create(nome="c", slug="c", organizacao=admin_user.organizacao)
    topico = TopicoDiscussao.objects.create(
        categoria=cat, titulo="t", slug="t", conteudo="x", autor=admin_user, publico_alvo=0
    )
    RespostaDiscussao.objects.create(topico=topico, autor=admin_user, conteudo="r")
    resp = client.get(
        reverse("dashboard:metrics-partial"),
        {
            "metricas": [
                "num_posts_feed_total",
                "num_posts_feed_recent",
                "num_topicos",
                "num_respostas",
            ]
        },
    )
    assert resp.status_code == 200
    content = resp.content.decode()
    assert "Posts (Total)" in content
    assert "Posts (24h)" in content
    assert "Tópicos" in content
    assert "Respostas" in content


def test_export_view_csv(monkeypatch, client, admin_user):
    client.force_login(admin_user)

    monkeypatch.setattr(
        DashboardMetricsService,
        "get_metrics",
        lambda *a, **kw: {"num_users": {"total": 1, "crescimento": 0.0}},
    )

    resp = client.get(reverse("dashboard:export"), {"formato": "csv"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == "text/csv"
    assert "attachment; filename=" in resp["Content-Disposition"]


def test_export_view_pdf(monkeypatch, client, admin_user):
    if weasyprint is None:
        pytest.skip("weasyprint não instalado")
    client.force_login(admin_user)

    monkeypatch.setattr(
        DashboardMetricsService,
        "get_metrics",
        lambda *a, **kw: {"num_users": {"total": 1, "crescimento": 0.0}},
    )
    monkeypatch.setattr("weasyprint.HTML.write_pdf", lambda self: b"pdf")

    resp = client.get(reverse("dashboard:export"), {"formato": "pdf"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"


def test_export_view_xlsx(monkeypatch, client, admin_user):
    client.force_login(admin_user)

    monkeypatch.setattr(
        DashboardMetricsService,
        "get_metrics",
        lambda *a, **kw: {"num_users": {"total": 1, "crescimento": 0.0}},
    )

    resp = client.get(reverse("dashboard:export"), {"formato": "xlsx"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(resp.content))
    ws = wb.active
    assert ws.title == "Métricas"
    assert ws.max_row == 2


def test_export_view_png(monkeypatch, client, admin_user, settings):
    client.force_login(admin_user)

    monkeypatch.setattr(
        DashboardMetricsService,
        "get_metrics",
        lambda *a, **kw: {"num_users": {"total": 1, "crescimento": 0.0}},
    )

    resp = client.get(reverse("dashboard:export"), {"formato": "png"})
    assert resp.status_code == 200
    assert resp["Content-Type"] == "image/png"
    filename = resp["Content-Disposition"].split("filename=")[1].strip('"')
    path = Path(settings.MEDIA_ROOT) / "dashboard_exports" / filename
    assert path.exists()


def test_export_view_permission(client, cliente_user):
    client.force_login(cliente_user)
    resp = client.get(reverse("dashboard:export"))
    assert resp.status_code == 403


def test_forbidden_dashboard_views(client, cliente_user):
    client.force_login(cliente_user)
    resp = client.get(reverse("dashboard:admin"))
    assert resp.status_code == 403
    resp = client.get(reverse("dashboard:root"))
    assert resp.status_code == 403


def test_config_save_and_apply(client, admin_user, gerente_user):
    client.force_login(admin_user)
    url = (
        reverse("dashboard:config-create")
        + "?periodo=mensal&escopo=organizacao&organizacao_id="
        + str(admin_user.organizacao_id)
        + "&metricas=num_users"
    )
    resp = client.post(url, {"nome": "cfg", "publico": True})
    assert resp.status_code == 302
    cfg = DashboardConfig.objects.get(nome="cfg")

    client.force_login(gerente_user)
    resp = client.get(reverse("dashboard:configs"))
    assert cfg in resp.context["object_list"]

    resp = client.get(reverse("dashboard:config-apply", args=[cfg.pk]))
    assert resp.status_code == 302
    assert "metricas=num_users" in resp["Location"]


def test_filter_save_and_apply(client, admin_user, gerente_user):
    client.force_login(admin_user)
    url = reverse("dashboard:filter-create") + "?periodo=mensal&escopo=auto&metricas=num_users"
    resp = client.post(url, {"nome": "f1", "publico": True})
    assert resp.status_code == 302
    filtro = DashboardFilter.objects.get(nome="f1")

    client.force_login(gerente_user)
    resp = client.get(reverse("dashboard:filters"))
    assert filtro in resp.context["object_list"]

    resp = client.get(reverse("dashboard:filter-apply", args=[filtro.pk]))
    assert resp.status_code == 302
    assert "metricas=num_users" in resp["Location"]
