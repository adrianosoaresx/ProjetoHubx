from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from dateutil.parser import parse
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import F, Q
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from ..models import Carteira, CentroCusto, ContaAssociado, LancamentoFinanceiro
from ..serializers import LancamentoFinanceiroSerializer
from .saldos import atribuir_carteiras_padrao


@dataclass
class ImportResult:
    """Resultado da importação de pagamentos."""

    preview: list[dict[str, Any]]
    errors: list[str]
    errors_file: str | None = None


class ImportadorPagamentos:
    """Serviço para importar pagamentos a partir de arquivos CSV ou XLSX."""

    REQUIRED = {
        "centro_custo_id",
        "tipo",
        "valor",
        "data_lancamento",
        "status",
    }
    OPTIONAL_ACCOUNT_FIELDS = {"conta_associado_id", "email"}

    def __init__(self, file_path: str, preview_limit: int = 5) -> None:
        self.file_path = Path(file_path)
        self.preview_limit = preview_limit

    # ────────────────────────────────────────────────────────────
    def _iter_rows(self) -> Iterable[dict[str, str]]:
        """Itera sobre as linhas do arquivo e gera dicionários de dados."""
        name = self.file_path.name.lower()
        if name.endswith(".csv"):
            with self.file_path.open(newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return
                headers = [h.lower() for h in reader.fieldnames]
                missing = self.REQUIRED - set(headers)
                if missing:
                    raise ValueError(_(f"Colunas faltantes: {', '.join(missing)}"))
                if not (set(headers) & self.OPTIONAL_ACCOUNT_FIELDS):
                    raise ValueError(_("Coluna conta_associado_id ou email obrigatória"))
                for row in reader:
                    yield {k.strip(): (v or "").strip() for k, v in row.items()}
        elif name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception as exc:  # pragma: no cover - openpyxl optional
                raise ValueError(_("openpyxl não disponível")) from exc
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            headers = next(rows)
            if not headers:
                return
            header_set = {str(h).lower() for h in headers}
            missing = self.REQUIRED - header_set
            if missing:
                raise ValueError(_(f"Colunas faltantes: {', '.join(missing)}"))
            if not (header_set & self.OPTIONAL_ACCOUNT_FIELDS):
                raise ValueError(_("Coluna conta_associado_id ou email obrigatória"))
            for values in rows:
                row = {str(h).strip(): str(v).strip() if v is not None else "" for h, v in zip(headers, values)}
                yield row
        else:
            raise ValueError(_("Formato não suportado"))

    # ────────────────────────────────────────────────────────────
    def preview(self) -> ImportResult:
        """Lê o arquivo e retorna uma prévia dos lançamentos."""
        preview: list[dict[str, Any]] = []
        errors: list[str] = []
        rejected: list[dict[str, str]] = []
        try:
            for idx, row in enumerate(self._iter_rows(), start=2):
                try:
                    record = self._convert_row(row)
                    if len(preview) < self.preview_limit:
                        preview.append(
                            {
                                "centro_custo": str(record["centro_custo"].id),
                                "conta_associado": (
                                    str(record["conta_associado"].id) if record.get("conta_associado") else None
                                ),
                                "tipo": record["tipo"],
                                "valor": str(record["valor"]),
                                "data_lancamento": record["data_lancamento"].isoformat(),
                                "data_vencimento": record["data_vencimento"].isoformat(),
                                "status": record["status"],
                                "descricao": record["descricao"],
                                "origem": record["origem"],
                            }
                        )
                except Exception as exc:
                    errors.append(f"Linha {idx}: {exc}")
                    rejected.append(row)
        except Exception as exc:
            errors.append(str(exc))
        errors_file = None
        if rejected:
            err_path = self.file_path.with_suffix(".errors.csv")
            with err_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=rejected[0].keys())
                writer.writeheader()
                writer.writerows(rejected)
            errors_file = str(err_path)
        return ImportResult(preview=preview, errors=errors, errors_file=errors_file)

    # ────────────────────────────────────────────────────────────
    def process(self, batch_size: int = 500) -> tuple[int, list[str]]:
        """Processa o arquivo criando lançamentos em lote."""
        errors: list[str] = []
        total = 0
        batch: list[dict[str, Any]] = []
        processed: set[tuple[str | None, str, str, Decimal, Any]] = set()

        def flush(chunk: list[dict[str, Any]]):
            to_create = []
            saldo_carteiras: dict[str, Decimal] = {}
            saldo_carteiras_contra: dict[str, Decimal] = {}
            carteiras_cache: dict[str, dict[Any, Carteira | None]] = {"centro": {}, "conta": {}}
            seen: set[tuple[str | None, str, str, Decimal, Any]] = set()
            validated: list[tuple[dict[str, Any], tuple[str | None, str, str, Decimal, Any]]] = []
            for item in chunk:
                payload = {
                    "centro_custo": str(item["centro_custo"].id),
                    "conta_associado": str(item["conta_associado"].id) if item.get("conta_associado") else None,
                    "tipo": item["tipo"],
                    "valor": str(item["valor"]),
                    "data_lancamento": item["data_lancamento"].isoformat(),
                    "data_vencimento": item["data_vencimento"].isoformat(),
                    "status": item["status"],
                    "descricao": item["descricao"],
                    "origem": item["origem"],
                }
                serializer = LancamentoFinanceiroSerializer(data=payload)
                if not serializer.is_valid():
                    errors.append(f"Dados inválidos na linha: {serializer.errors}")
                    continue
                data = serializer.validated_data
                atribuir_carteiras_padrao(data, cache=carteiras_cache)
                conta = data.get("conta_associado")
                key = (
                    str(data["centro_custo"].id),
                    str(conta.id) if conta else None,
                    data["tipo"],
                    data["valor"],
                    data["data_lancamento"],
                )
                validated.append((data, key))

            if validated:
                q = Q()
                for _, key in validated:
                    q |= Q(
                        centro_custo_id=key[0],
                        conta_associado_id=key[1],
                        tipo=key[2],
                        valor=key[3],
                        data_lancamento=key[4],
                    )
                existing = {
                    (
                        str(cc),
                        str(ca) if ca else None,
                        tipo,
                        valor,
                        dl,
                    )
                    for cc, ca, tipo, valor, dl in LancamentoFinanceiro.objects.filter(q).values_list(
                        "centro_custo_id",
                        "conta_associado_id",
                        "tipo",
                        "valor",
                        "data_lancamento",
                    )
                }
            else:
                existing = set()
            existing.update(processed)

            for data, key in validated:
                if key in seen or key in existing:
                    errors.append("Lançamento duplicado")
                    continue
                seen.add(key)
                processed.add(key)
                if data["status"] == LancamentoFinanceiro.Status.PAGO:
                    valor_lanc = data["valor"]
                    carteira = data.get("carteira")
                    if carteira:
                        cid = str(carteira.id)
                        saldo_carteiras[cid] = saldo_carteiras.get(cid, Decimal("0")) + valor_lanc
                    conta = data.get("conta_associado")
                    carteira_contra = data.get("carteira_contraparte")
                    if conta and carteira_contra:
                        aid = str(carteira_contra.id)
                        saldo_carteiras_contra[aid] = saldo_carteiras_contra.get(aid, Decimal("0")) + valor_lanc
                to_create.append(LancamentoFinanceiro(**data))
            if to_create:
                LancamentoFinanceiro.objects.bulk_create(to_create)
                nonlocal total
                total += len(to_create)
            for cid, inc in saldo_carteiras.items():
                Carteira.objects.filter(pk=cid).update(saldo=F("saldo") + inc)
            for aid, inc in saldo_carteiras_contra.items():
                Carteira.objects.filter(pk=aid).update(saldo=F("saldo") + inc)

        try:
            for idx, row in enumerate(self._iter_rows(), start=2):
                try:
                    batch.append(self._convert_row(row))
                except Exception as exc:
                    errors.append(f"Linha {idx}: {exc}")
                    continue
                if len(batch) >= batch_size:
                    with transaction.atomic():
                        flush(batch)
                    batch = []
            if batch:
                with transaction.atomic():
                    flush(batch)
        except Exception:
            raise
        return total, errors

    # ────────────────────────────────────────────────────────────
    def _convert_row(self, row: dict[str, str]) -> dict[str, Any]:
        """Converte uma linha do arquivo em dados validados."""
        centro = CentroCusto.objects.get(pk=row["centro_custo_id"])
        conta: ContaAssociado | None = None
        cid = row.get("conta_associado_id")
        email = row.get("email")
        if cid:
            conta = ContaAssociado.objects.filter(pk=cid).first()
        elif email:
            User = get_user_model()
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                raise ValidationError(_("Usuário com e-mail não encontrado"))
            conta = ContaAssociado.objects.filter(user=user).first()
        if not conta:
            raise ValueError(_("Conta do associado não encontrada"))
        data_lanc = parse(row["data_lancamento"])
        if not data_lanc.tzinfo:
            data_lanc = timezone.make_aware(data_lanc)
        data_vencimento = row.get("data_vencimento")
        if data_vencimento:
            venc = parse(data_vencimento)
            if not venc.tzinfo:
                venc = timezone.make_aware(venc)
            if venc < data_lanc:
                raise ValueError(_("data_vencimento anterior a data_lancamento"))
        else:
            venc = data_lanc
        valor = Decimal(row["valor"])
        if valor < 0 and row["tipo"] != LancamentoFinanceiro.Tipo.DESPESA:
            raise ValueError(_("Valor negativo não permitido para este tipo"))
        return {
            "centro_custo": centro,
            "conta_associado": conta,
            "tipo": row["tipo"],
            "valor": valor,
            "data_lancamento": data_lanc,
            "data_vencimento": venc,
            "status": row["status"],
            "descricao": row.get("descricao", ""),
            "origem": LancamentoFinanceiro.Origem.IMPORTACAO,
        }
